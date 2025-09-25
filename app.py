import json, os, re, secrets
from functools import wraps
from datetime import datetime, timedelta, timezone
from pathlib import Path
from collections import Counter

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, abort, make_response, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv

# Export/Import helpers
import csv
from io import StringIO, BytesIO

# XLSX support: use openpyxl if available; otherwise fall back to CSV
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None
    
# --------------------------------------------------------------------------------------
# Paths & env
# --------------------------------------------------------------------------------------
BASE_DIR     = Path(__file__).parent
DATA_DIR     = BASE_DIR / "data"
INV_FILE     = DATA_DIR / "inventory.json"
USERS_FILE   = DATA_DIR / "users.json"
RESET_FILE   = DATA_DIR / "reset_tokens.json"
CONTACT_FILE = DATA_DIR / "contact_messages.json"
ENV_FILE     = BASE_DIR / ".env"

DATA_DIR.mkdir(exist_ok=True)
load_dotenv(ENV_FILE)

# --------------------------------------------------------------------------------------
# App
# --------------------------------------------------------------------------------------
app = Flask(__name__, static_folder="static", template_folder="templates")

def env_set_or_update(key: str, value: str):
    """Replace or append a key in .env safely and update process env."""
    ENV_FILE.touch(exist_ok=True)
    lines = ENV_FILE.read_text(encoding="utf-8").splitlines()
    pattern = re.compile(rf"^{re.escape(key)}=", re.I)
    for i, line in enumerate(lines):
        if pattern.match(line):
            lines[i] = f"{key}={value}"
            break
    else:
        lines.append(f"{key}={value}")
    ENV_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")
    os.environ[key] = value

# Ensure SECRET_KEY exists
if not os.getenv("SECRET_KEY"):
    env_set_or_update("SECRET_KEY", secrets.token_urlsafe(32))
app.secret_key = os.getenv("SECRET_KEY")

# Cookie settings
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE", "false").lower()=="true",
)

# --------------------------------------------------------------------------------------
# Helpers (IO)
# --------------------------------------------------------------------------------------
def read_json(p: Path, default):
    if not p.exists():
        p.write_text(json.dumps(default, ensure_ascii=False, indent=2), encoding="utf-8")
        return default
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default

def write_json(p: Path, data):
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# --------------------------------------------------------------------------------------
# Auth helpers
# --------------------------------------------------------------------------------------
def get_admin_creds():
    return os.getenv("ADMIN_USERNAME", "admin"), os.getenv("ADMIN_PASSWORD_HASH", "")

def is_authed():
    return bool(session.get("user"))

def current_user():
    return session.get("user") or {}

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_authed():
            abort(401)
        return fn(*args, **kwargs)
    return wrapper

# --------------------------------------------------------------------------------------
# Contact / notifications storage
# --------------------------------------------------------------------------------------
def read_messages():
    return read_json(CONTACT_FILE, [])

def write_messages(msgs):
    write_json(CONTACT_FILE, msgs)

# --------------------------------------------------------------------------------------
# Configs
# --------------------------------------------------------------------------------------
REGISTER_OPEN   = os.getenv("REGISTER_OPEN", "false").lower() == "true"
INVITE_CODE     = os.getenv("REGISTER_INVITE_CODE", "").strip()
USERNAME_RE     = re.compile(r"^[a-zA-Z0-9_.-]{3,32}$")
RESET_TTL       = int(os.getenv("RESET_TOKEN_TTL_MIN", "20"))
RESET_DEV_SHOW  = os.getenv("RESET_DEV_SHOW_TOKEN", "false").lower() == "true"
DEFAULT_ROLE    = os.getenv("DEFAULT_ROLE", "user")

# Navbar badge (unread contact messages)
@app.context_processor
def inject_notif_count():
    try:
        msgs = read_messages()
    except Exception:
        msgs = []
    unread = sum(1 for m in msgs if not m.get("read"))
    return dict(notif_count=unread)

# --------------------------------------------------------------------------------------
# Reset token helpers
# --------------------------------------------------------------------------------------
def read_tokens():
    return read_json(RESET_FILE, [])

def write_tokens(toks):
    write_json(RESET_FILE, toks)

def prune_tokens(toks):
    now = datetime.now(timezone.utc)
    out = []
    for t in toks:
        try:
            exp = datetime.fromisoformat(t.get("exp"))
            if exp > now:
                out.append(t)
        except Exception:
            continue
    return out

# --------------------------------------------------------------------------------------
# Users
# --------------------------------------------------------------------------------------
def username_exists(username: str) -> bool:
    env_key = f"USER_{username.upper()}_HASH"
    if os.getenv(env_key):
        return True
    users = read_json(USERS_FILE, [])
    return any(u.get("username") == username for u in users)

# --------------------------------------------------------------------------------------
# Routes: auth
# --------------------------------------------------------------------------------------
@app.get("/")
def home():
    if not is_authed():
        return redirect(url_for("login"))
    return redirect(url_for("dashboard"))

@app.get("/login")
def login():
    if is_authed():
        return redirect(url_for("dashboard"))
    return render_template("login.html", REGISTER_OPEN=REGISTER_OPEN)

@app.post("/login")
def do_login():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    admin_user, admin_hash = get_admin_creds()

    # Admin path
    if username == admin_user and admin_hash and check_password_hash(admin_hash, password):
        session["user"] = {"username": username, "role": "admin"}
        return redirect(url_for("dashboard"))

    # Normal users from env
    env_key  = f"USER_{username.upper()}_HASH"
    user_hash = os.getenv(env_key)
    if user_hash and check_password_hash(user_hash, password):
        users = read_json(USERS_FILE, [])
        role = next((u.get("role", DEFAULT_ROLE) for u in users if u.get("username") == username), DEFAULT_ROLE)
        session["user"] = {"username": username, "role": role}
        return redirect(url_for("dashboard"))

    return render_template("login.html", error="Geçersiz kullanıcı adı veya şifre", REGISTER_OPEN=REGISTER_OPEN)

@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# --------------------------------------------------------------------------------------
# Routes: register
# --------------------------------------------------------------------------------------
@app.get("/register")
def register_page():
    if not REGISTER_OPEN:
        abort(404)
    if is_authed():
        return redirect(url_for("dashboard"))
    return render_template("register.html", INVITE_CODE=INVITE_CODE)

@app.post("/register")
def do_register():
    if not REGISTER_OPEN:
        abort(404)
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    confirm  = request.form.get("confirm") or ""
    invite   = (request.form.get("invite") or "").strip()

    if not USERNAME_RE.match(username):
        return render_template("register.html", error="Kullanıcı adı 3–32 karakter olmalı ve harf/rakam/._- içermelidir.", INVITE_CODE=INVITE_CODE)
    if username_exists(username):
        return render_template("register.html", error="Bu kullanıcı adı zaten kullanılıyor.", INVITE_CODE=INVITE_CODE)
    if password != confirm:
        return render_template("register.html", error="Şifreler eşleşmiyor.", INVITE_CODE=INVITE_CODE)
    if len(password) < 8 or not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        return render_template("register.html", error="Şifre en az 8 karakter, harf ve rakam içermelidir.", INVITE_CODE=INVITE_CODE)
    if INVITE_CODE and invite != INVITE_CODE:
        return render_template("register.html", error="Geçersiz davet kodu.", INVITE_CODE=INVITE_CODE)

    env_set_or_update(f"USER_{username.upper()}_HASH", generate_password_hash(password))
    users = read_json(USERS_FILE, [])
    users.append({"username": username, "role": DEFAULT_ROLE})
    write_json(USERS_FILE, users)
    return redirect(url_for("login"))

# --------------------------------------------------------------------------------------
# Routes: contact / notifications
# --------------------------------------------------------------------------------------
@app.get("/contact")
def contact_page():
    return render_template("contact.html", sent=False)

@app.post("/contact")
def contact_submit():
    name        = (request.form.get("name") or "").strip()
    email       = (request.form.get("email") or "").strip()
    subject     = (request.form.get("subject") or "").strip()
    department  = (request.form.get("department") or "").strip()
    message     = (request.form.get("message") or "").strip()
    consent     = request.form.get("consent") == "on"

    errors = []
    if len(name) < 2: errors.append("İsim en az 2 karakter olmalı")
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email or ""): errors.append("Geçerli bir e-posta giriniz")
    if len(subject) < 3: errors.append("Konu en az 3 karakter olmalı")
    if len(message) < 10: errors.append("Mesaj en az 10 karakter olmalı")
    if not consent: errors.append("KVKK/ileti onayı gerekli")

    if errors:
        return render_template("contact.html", sent=False, errors=errors,
                               name=name, email=email, subject=subject, department=department, message=message)

    msgs = read_messages()
    msgs.append({
        "id": secrets.token_hex(8),
        "ts": datetime.now(timezone.utc).isoformat(),
        "name": name, "email": email, "subject": subject,
        "department": department, "message": message,
        "read": False
    })
    write_messages(msgs)
    return render_template("contact.html", sent=True)

@app.get("/notifications")
@login_required
def notifications_page():
    msgs = sorted(read_messages(), key=lambda m: m.get("ts",""), reverse=True)
    return render_template("notifications.html", messages=msgs)

@app.post("/api/notifications/<msg_id>/read")
@login_required
def api_mark_read(msg_id):
    payload = request.json or {}
    to_val = bool(payload.get("read", True))
    msgs = read_messages()
    updated = False
    for m in msgs:
        if m.get("id") == msg_id:
            m["read"] = to_val
            updated = True
            break
    if updated:
        write_messages(msgs)
    unread = sum(1 for m in msgs if not m.get("read"))
    return jsonify({"ok": updated, "unread": unread})

@app.post("/api/notifications/read-all")
@login_required
def api_mark_all_read():
    msgs = read_messages()
    for m in msgs:
        m["read"] = True
    write_messages(msgs)
    return jsonify({"ok": True, "unread": 0})

# --------------------------------------------------------------------------------------
# Routes: forgot / reset
# --------------------------------------------------------------------------------------
@app.get("/forgot")
def forgot_page():
    if is_authed():
        return redirect(url_for("dashboard"))
    return render_template("forgot.html", sent=False, dev=RESET_DEV_SHOW)

@app.post("/forgot")
def forgot_submit():
    username = (request.form.get("username") or "").strip()
    admin_user, _ = get_admin_creds()
    exists = (username == admin_user) or username_exists(username)

    reset_link = None
    if exists:
        toks = prune_tokens(read_tokens())
        token = secrets.token_urlsafe(32)
        exp = (datetime.now(timezone.utc) + timedelta(minutes=RESET_TTL)).isoformat()
        toks.append({"username": username, "token": token, "exp": exp})
        write_tokens(toks)
        if RESET_DEV_SHOW:
            reset_link = url_for("reset_page", token=token, _external=True)
            print("[DEV] Password reset link:", reset_link)
    return render_template("forgot.html", sent=True, dev=RESET_DEV_SHOW, link=reset_link)

@app.get("/reset")
def reset_page():
    token = request.args.get("token", "")
    toks = prune_tokens(read_tokens())
    match = next((t for t in toks if t.get("token") == token), None)
    if not match:
        return render_template("reset.html", invalid=True)
    return render_template("reset.html", invalid=False, token=token)

@app.post("/reset")
def reset_submit():
    token   = request.form.get("token", "")
    new     = request.form.get("new", "")
    confirm = request.form.get("confirm", "")

    if new != confirm or len(new) < 8 or not re.search(r"[A-Za-z]", new) or not re.search(r"\d", new):
        return render_template("reset.html", invalid=False, token=token,
                               error="Yeni şifre en az 8 karakter, harf ve rakam içermeli ve eşleşmeli")

    toks  = prune_tokens(read_tokens())
    match = next((t for t in toks if t.get("token") == token), None)
    if not match:
        return render_template("reset.html", invalid=True)

    username = match.get("username")
    admin_user, _ = get_admin_creds()
    if username == admin_user:
        env_set_or_update("ADMIN_PASSWORD_HASH", generate_password_hash(new))
    else:
        env_set_or_update(f"USER_{username.upper()}_HASH", generate_password_hash(new))

    toks = [t for t in toks if t.get("token") != token]
    write_tokens(toks)
    return redirect(url_for("login"))

# --------------------------------------------------------------------------------------
# Routes: dashboard
# --------------------------------------------------------------------------------------
@app.get("/dashboard")
@login_required
def dashboard():
    inv = read_json(INV_FILE, [])
    total = len(inv)

    def norm(v, fallback):
        s = (v or "").strip() if isinstance(v, str) else v
        return s or fallback

    by_status = Counter(norm(r.get("status"), "Bilinmiyor") for r in inv)
    by_os     = Counter(norm(r.get("os") or r.get("operating_system"), "Diğer") for r in inv)

    return render_template(
        "dashboard.html",
        user=current_user(),
        total=total,
        by_status=dict(by_status),
        by_os=dict(by_os),
    )

# --------------------------------------------------------------------------------------
# Routes: inventory pages & APIs
# --------------------------------------------------------------------------------------
@app.get("/inventory")
@login_required
def inventory_page():
    return render_template("inventory.html", user=current_user())

# ---- Inventory GET (single definition; includes legacy key normalization) ----
@app.get("/api/inventory")
@login_required
def api_inventory_list():
    data = read_json(INV_FILE, [])

    # Normalize legacy keys -> new schema (owner/department/model)
    def _normalize(row: dict) -> dict:
        r = dict(row)
        if not r.get("owner") and r.get("user"):
            r["owner"] = r["user"]
        if not r.get("department") and r.get("location"):
            r["department"] = r["location"]
        if not r.get("model") and r.get("name"):
            r["model"] = r["name"]
        return r
    data = [_normalize(x) for x in data]

    # Optional lightweight filtering
    q = (request.args.get("q") or "").strip().lower()
    if q:
        def _hit(row):
            blob = json.dumps(row, ensure_ascii=False).lower()
            return q in blob
        data = [r for r in data if _hit(r)]

    # Optional pagination
    try:
        offset = max(0, int(request.args.get("offset", 0)))
        limit  = min(500, max(1, int(request.args.get("limit", 1000))))
    except ValueError:
        offset, limit = 0, 1000
    return jsonify(data[offset:offset+limit])

# ---- Inventory POST (aligned to new schema; accepts legacy keys for safety) ----
@app.post("/api/inventory")
@login_required
def api_inventory_add():
    payload = request.json or {}
    inv = read_json(INV_FILE, [])

    item = {
        "id": payload.get("id") if payload.get("id") else _new_item_id(inv),
        "owner": _sanitize_str(payload.get("owner"), 120) or _sanitize_str(payload.get("user"), 120),
        "department": _sanitize_str(payload.get("department"), 120) or _sanitize_str(payload.get("location"), 120),
        "model": _sanitize_str(payload.get("model"), 120) or _sanitize_str(payload.get("name"), 120),
        "ip": _sanitize_str(payload.get("ip"), 120),
        "os": _sanitize_str(payload.get("os"), 80),
        "status": _sanitize_str(payload.get("status"), 40) or "Bilinmiyor",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    inv.append(item)
    write_json(INV_FILE, inv)
    return jsonify({"ok": True, "item": item})

# ---- Inventory PUT ----
@app.put("/api/inventory/<item_id>")
@login_required
def api_inventory_update(item_id):
    payload = request.json or {}
    inv = read_json(INV_FILE, [])
    for i, row in enumerate(inv):
        if str(row.get("id")) == str(item_id):
            row.update({
                "owner": _sanitize_str(payload.get("owner", row.get("owner")), 120) or _sanitize_str(payload.get("user", row.get("owner")), 120),
                "department": _sanitize_str(payload.get("department", row.get("department")), 120) or _sanitize_str(payload.get("location", row.get("department")), 120),
                "model": _sanitize_str(payload.get("model", row.get("model")), 120) or _sanitize_str(payload.get("name", row.get("model")), 120),
                "ip": _sanitize_str(payload.get("ip", row.get("ip")), 120),
                "os": _sanitize_str(payload.get("os", row.get("os")), 80),
                "status": _sanitize_str(payload.get("status", row.get("status")), 40),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
            inv[i] = row
            write_json(INV_FILE, inv)
            return jsonify({"ok": True, "item": row})
    return jsonify({"ok": False, "error": "Not found"}), 404

# ---- Inventory DELETE ----
@app.delete("/api/inventory/<item_id>")
@login_required
def api_inventory_delete(item_id):
    inv = read_json(INV_FILE, [])
    new_inv = [r for r in inv if str(r.get("id")) != str(item_id)]
    write_json(INV_FILE, new_inv)
    return jsonify({"ok": True, "deleted": str(item_id)})

# Helpers used by inventory endpoints

def _sanitize_str(s: str, max_len=200):
    if s is None: return ""
    s = str(s).strip()
    return s[:max_len]

def _new_item_id(existing:list):
    # integer-like auto id
    ints = [int(r.get("id")) for r in existing if str(r.get("id","")).isdigit()]
    return (max(ints) + 1) if ints else 1

# --------------------------------------------------------------------------------------
# Routes: Export / Import
# --------------------------------------------------------------------------------------
REQUIRED_COLUMNS = ["id","owner","department","model","ip","os","status"]

def _rows_to_csv(rows:list) -> bytes:
    buf = StringIO()
    if not rows:
        writer = csv.DictWriter(buf, fieldnames=REQUIRED_COLUMNS)
        writer.writeheader()
    else:
        cols = list(dict.fromkeys(REQUIRED_COLUMNS + list(rows[0].keys())))
        writer = csv.DictWriter(buf, fieldnames=cols)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    return buf.getvalue().encode("utf-8-sig")  # Excel-friendly BOM

# Flexible export (fmt=csv|json). Keep legacy callers working.
@app.route("/api/inventory/export", methods=["GET","POST"])
@login_required
def api_inventory_export():
    fmt = (request.args.get("fmt") or request.form.get("fmt") or "csv").lower()
    data = read_json(INV_FILE, [])
    filename = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{fmt}"

    if fmt == "json":
        payload = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        return send_file(BytesIO(payload), as_attachment=True,
                         download_name=filename, mimetype="application/json")

    # default CSV
    payload = _rows_to_csv(data)
    return send_file(BytesIO(payload), as_attachment=True,
                     download_name=filename, mimetype="text/csv")

@app.get("/api/inventory/export.csv")
@login_required
def api_inventory_export_csv():
    data = read_json(INV_FILE, [])
    payload = _rows_to_csv(data)
    return send_file(BytesIO(payload), as_attachment=True,
                     download_name="inventory.csv", mimetype="text/csv")

@app.get("/api/inventory/export.xlsx")
@login_required
def api_inventory_export_xlsx():
    data = read_json(INV_FILE, [])
    if Workbook is None:
        return redirect(url_for("api_inventory_export_csv"))

    cols = list(dict.fromkeys(REQUIRED_COLUMNS + (list(data[0].keys()) if data else [])))
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(cols)
    for r in data:
        ws.append([r.get(c, "") for c in cols])
    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(
        bio, as_attachment=True, download_name="inventory.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/inventory/sample.csv")
@login_required
def api_inventory_sample_csv():
    sample = [{
        "id":"A-2001","owner":"Ali Veli","department":"Yazılım",
        "model":"Dell Latitude 5520","ip":"192.168.1.50",
        "os":"Windows 11 Pro","status":"Aktif"
    }]
    payload = _rows_to_csv(sample)
    return send_file(BytesIO(payload), as_attachment=True,
                     download_name="sample.csv", mimetype="text/csv")

@app.get("/api/inventory/sample.xlsx")
@login_required
def api_inventory_sample_xlsx():
    if Workbook is None:
        return redirect(url_for("api_inventory_sample_csv"))
    sample = [{
        "id":"A-2001","owner":"Ali Veli","department":"Yazılım",
        "model":"Dell Latitude 5520","ip":"192.168.1.50",
        "os":"Windows 11 Pro","status":"Aktif"
    }]
    cols = list(sample[0].keys())
    wb = Workbook(); ws = wb.active; ws.title = "Sample"
    ws.append(cols)
    for r in sample:
        ws.append([r.get(c,"") for c in cols])
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(
        bio, as_attachment=True, download_name="sample.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/api/inventory/import", methods=["POST"])
@login_required
def api_inventory_import():
    """
    Accept CSV or XLSX upload (multipart/form-data, field name: 'file').
    Required columns (case-insensitive): id, owner, department, model, ip, os, status

    - If an incoming 'id' matches existing, row is UPDATED
    - Otherwise it is APPENDED
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "file required"}), 400

    f = request.files["file"]
    filename = (f.filename or "").lower().strip()

    # Helper: normalize dict keys to lower and strip
    def _norm_keys(d: dict) -> dict:
        out = {}
        for k, v in (d or {}).items():
            if k is None:
                continue
            key = str(k).strip().lower()
            val = "" if v is None else str(v).strip()
            out[key] = val
        return out

    # Read rows into list[dict] with lowercased keys
    rows = []
    header_lower = []

    if filename.endswith(".xlsx"):
        if not load_workbook:
            return jsonify({"ok": False, "error": "XLSX import requires openpyxl. Install it or upload CSV."}), 400
        bio = BytesIO(f.read())
        wb = load_workbook(bio, read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return jsonify({"ok": False, "error": "empty file"}), 400

        # header: first row
        header = [(str(x).strip() if x is not None else "") for x in raw_rows[0]]
        header_lower = [h.lower() for h in header]
        # required check
        if not set(REQUIRED_COLUMNS).issubset(set(header_lower)):
            return jsonify({"ok": False, "error": f"missing columns, need: {', '.join(REQUIRED_COLUMNS)}"}), 400

        for r in raw_rows[1:]:
            if r is None:
                continue
            # skip fully empty lines
            if all((c is None or str(c).strip() == "") for c in r):
                continue
            rec = {}
            for i, cell in enumerate(r):
                if i >= len(header_lower):
                    break
                key = header_lower[i]
                if not key:
                    continue
                rec[key] = "" if cell is None else str(cell).strip()
            rows.append(_norm_keys(rec))

    else:
        # Default: CSV
        # decode safely
        data = f.read()
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1")

        # Manual CSV parse to control header normalization
        rdr = csv.reader(StringIO(text))
        raw = list(rdr)
        if not raw:
            return jsonify({"ok": False, "error": "empty file"}), 400

        header = [str(h or "").strip() for h in raw[0]]
        header_lower = [h.lower() for h in header]
        if not set(REQUIRED_COLUMNS).issubset(set(header_lower)):
            return jsonify({"ok": False, "error": f"missing columns, need: {', '.join(REQUIRED_COLUMNS)}"}), 400

        for line in raw[1:]:
            if not any(x.strip() for x in map(lambda x: str(x or ""), line)):
                continue
            rec = {}
            for i, h in enumerate(header_lower):
                rec[h] = str(line[i]).strip() if i < len(line) and line[i] is not None else ""
            rows.append(_norm_keys(rec))

    # Upsert into inventory.json
    inv = read_json(INV_FILE, [])
    index = {str(r.get("id")): i for i, r in enumerate(inv)}

    updated, inserted = 0, 0
    for row in rows:
        # keep only expected keys + preserve others if present
        clean = {
            "id": row.get("id", ""),
            "owner": row.get("owner", ""),
            "department": row.get("department", ""),
            "model": row.get("model", ""),
            "ip": row.get("ip", ""),
            "os": row.get("os", ""),
            "status": row.get("status", "") or "Bilinmiyor",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }

        rid = str(clean.get("id") or "").strip()
        if rid and rid in index:
            # update existing
            inv[index[rid]].update(clean)
            updated += 1
        else:
            # if id empty, auto-assign integer-like id
            if not rid:
                rid = str(_new_item_id(inv))
                clean["id"] = rid
            inv.append(clean)
            index[rid] = len(inv) - 1
            inserted += 1

    write_json(INV_FILE, inv)
    return jsonify({"ok": True, "updated": updated, "inserted": inserted, "total": updated + inserted})

# --------------------------------------------------------------------------------------
# Routes: admin
# --------------------------------------------------------------------------------------
@app.get("/admin")
@login_required
def admin_page():
    if current_user().get("role") != "admin":
        abort(403)
    users = read_json(USERS_FILE, [])
    return render_template("admin.html", user=current_user(), users=users)

@app.post("/api/admin/user-hash")
@login_required
def api_admin_set_user_hash():
    if current_user().get("role") != "admin":
        abort(403)
    payload = request.json or {}
    username     = (payload.get("username") or "").strip()
    new_password = payload.get("password") or ""
    if not username or not new_password:
        return jsonify({"ok": False, "error": "username and password required"}), 400

    env_set_or_update(f"USER_{username.upper()}_HASH", generate_password_hash(new_password))

    users = read_json(USERS_FILE, [])
    existing = next((u for u in users if u.get("username") == username), None)
    if not existing:
        users.append({"username": username, "role": DEFAULT_ROLE})
    role = payload.get("role")
    if role:
        for u in users:
            if u.get("username") == username:
                u["role"] = role
    write_json(USERS_FILE, users)
    return jsonify({"ok": True})

@app.post("/api/change-password")
@login_required
def api_change_password():
    username = current_user().get("username")
    old = (request.json or {}).get("old", "")
    new = (request.json or {}).get("new", "")

    if not new:
        return jsonify({"ok": False, "error": "new password required"}), 400

    admin_user, admin_hash = get_admin_creds()
    if username == admin_user:
        if not admin_hash or not check_password_hash(admin_hash, old):
            return jsonify({"ok": False, "error": "old password incorrect"}), 400
        env_set_or_update("ADMIN_PASSWORD_HASH", generate_password_hash(new))
        return jsonify({"ok": True})
    else:
        key = f"USER_{username.upper()}_HASH"
        cur_hash = os.getenv(key)
        if not cur_hash or not check_password_hash(cur_hash, old):
            return jsonify({"ok": False, "error": "old password incorrect"}), 400
        env_set_or_update(key, generate_password_hash(new))
        return jsonify({"ok": True})

# --------------------------------------------------------------------------------------
# Contact inbox APIs (dashboard widget) + delete (single/bulk)
# --------------------------------------------------------------------------------------
@app.get("/api/contact/unread-count")
@login_required
def api_contact_unread_count():
    """Return unread count for the top-right badge on the widget/card."""
    try:
        msgs = read_messages()
    except Exception:
        msgs = []
    unread = sum(1 for m in msgs if not m.get("read"))
    return jsonify({"unread": unread})

@app.get("/api/contact/latest")
@login_required
def api_contact_latest():
    """
    Return latest N contact messages for the dashboard widget.
    Query: ?limit=5 (default 5)
    """
    try:
        limit = max(1, min(50, int(request.args.get("limit", 5))))
    except Exception:
        limit = 5

    msgs = read_messages()

    # sort by timestamp (iso string), newest first
    def _ts(m): return m.get("ts") or m.get("created_at") or ""
    latest = sorted(msgs, key=_ts, reverse=True)[:limit]

    # trim fields for the widget
    out = []
    for i, m in enumerate(latest):
        out.append({
            "id": m.get("id", i),
            "name": m.get("name") or m.get("sender") or "",
            "email": m.get("email") or "",
            "subject": m.get("subject") or "(konu yok)",
            "message": m.get("message") or "",
            "created_at": m.get("ts") or m.get("created_at") or "",
            "read": bool(m.get("read")),
        })
    return jsonify(out)

@app.post("/api/contact/<mid>/read")
@login_required
def api_contact_mark_read(mid):
    msgs = read_messages()
    updated = False
    for m in msgs:
        if str(m.get("id")) == str(mid):
            m["read"] = True
            updated = True
            break
    if updated:
        write_messages(msgs)
    unread = sum(1 for m in msgs if not m.get("read"))
    return jsonify({"ok": updated, "unread": unread})

@app.post("/api/contact/<mid>/unread")
@login_required
def api_contact_mark_unread(mid):
    msgs = read_messages()
    updated = False
    for m in msgs:
        if str(m.get("id")) == str(mid):
            m["read"] = False
            updated = True
            break
    if updated:
        write_messages(msgs)
    unread = sum(1 for m in msgs if not m.get("read"))
    return jsonify({"ok": updated, "unread": unread})

@app.delete("/api/contact/<mid>")
@login_required
def api_contact_delete_one(mid):
    msgs = read_messages()
    before = len(msgs)
    msgs = [m for m in msgs if str(m.get("id")) != str(mid)]
    deleted = before - len(msgs)
    if deleted:
        write_messages(msgs)
    unread = sum(1 for m in msgs if not m.get("read"))
    return jsonify({"ok": bool(deleted), "deleted": deleted, "remaining": len(msgs), "unread": unread})

@app.post("/api/contact/delete-bulk")
@login_required
def api_contact_delete_bulk():
    """
    Delete multiple contact messages.
    Body (JSON):
      { "ids": ["id1","id2", ...] }  -> delete selected
      { "all": true }                -> delete ALL
    """
    payload = request.get_json(silent=True) or {}
    msgs = read_messages()
    before = len(msgs)

    if payload.get("all"):
        msgs = []
    else:
        ids = set(str(x) for x in (payload.get("ids") or []))
        if not ids:
            return jsonify({"ok": False, "error": "ids or all required"}), 400
        msgs = [m for m in msgs if str(m.get("id")) not in ids]

    deleted = before - len(msgs)
    write_messages(msgs)
    unread = sum(1 for m in msgs if not m.get("read"))
    return jsonify({"ok": True, "deleted": deleted, "remaining": len(msgs), "unread": unread})

# --------------------------------------------------------------------------------------
# Error handlers + security headers
# --------------------------------------------------------------------------------------
@app.errorhandler(401)
def _401(e):
    return redirect(url_for("login"))

@app.errorhandler(403)
def _403(e):
    return render_template("base.html", title="Erişim engellendi"), 403

@app.after_request
def add_security_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    # Basic CSP that works with your current assets (Chart.js via CDN allowed)
    csp = (
        "default-src 'self'; "
        "script-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "style-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "img-src 'self' data:; "
        "font-src https://cdn.jsdelivr.net 'self'; "
        "connect-src 'self'; "
        "frame-ancestors 'self';"
    )
    resp.headers.setdefault("Content-Security-Policy", csp)
    return resp

# --------------------------------------------------------------------------------------
# Entrypoint
# --------------------------------------------------------------------------------------
if __name__ == "__main__":
    host = os.getenv("FLASK_HOST", "127.0.0.1")
    port = int(os.getenv("FLASK_PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "true").lower() == "true"
    app.run(host=host, port=port, debug=debug)
